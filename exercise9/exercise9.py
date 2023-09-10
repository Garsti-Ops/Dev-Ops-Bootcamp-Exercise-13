import openpyxl
from openpyxl.workbook import Workbook

from employee import Employee

employees_file = openpyxl.load_workbook("../employees.xlsx")
sheet1 = employees_file["Sheet1"]

cells = sheet1["A1:D1"]
employees = []

for row in range(2, sheet1.max_row + 1):
    e = Employee(sheet1.cell(row, 1).value,
                 sheet1.cell(row, 2).value,
                 sheet1.cell(row, 3).value,
                 sheet1.cell(row, 4).value)

    employees.append(e)

for i in range(len(employees) - 1):
    for j in range(0, len(employees) - i - 1):
        obj = employees[j]
        obj2 = employees[j + 1]

        if obj.years_of_experience < obj2.years_of_experience:
            employees[j + 1] = obj
            employees[j] = obj2

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

ws.cell(1, 1, "Name")
ws.cell(1, 2, "Years of Experience")
ws.cell(1, 3, "Job Title")
ws.cell(1, 4, "Date of Birth")

row = 2
for employee in employees:
    cell = 1
    while cell < 4:
        ws.cell(row, cell, employee.get_values_as_array()[cell - 1])
        cell += 1
    row += 1

wb.save("employees_sorted.xlsx")
